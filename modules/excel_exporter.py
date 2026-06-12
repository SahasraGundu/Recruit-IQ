# modules/excel_exporter.py
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Colors ───────────────────────────────────────────────────────────────────
INDIGO      = "4F46E5"
VIOLET      = "7C3AED"
LIGHT_BG    = "EEF2FF"
SUCCESS     = "059669"
DANGER      = "DC2626"
WARNING     = "D97706"
WHITE       = "FFFFFF"
DARK        = "1E1B4B"
GRAY        = "6B7280"
LIGHT_GRAY  = "F8F9FF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="111827", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    thin = Side(style="thin", color="C7D2FE")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _set_row(ws, row_num, values, fill_hex, bold=False,
             font_color="111827", font_size=10, h_align="left"):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill      = _fill(fill_hex)
        cell.font      = _font(bold=bold, color=font_color, size=font_size)
        cell.alignment = _align(h=h_align, wrap=True)
        cell.border    = _border()


def generate_excel_report(drive: dict, candidates: list, jd_parsed: dict) -> bytes:
    """Generate full Excel workbook. Returns bytes."""

    wb = openpyxl.Workbook()

    # ── Remove default sheet ──────────────────────────────────────────────────
    wb.remove(wb.active)

    drive_name = drive.get("drive_name", "Drive")
    company    = drive.get("company",    "Company")
    role       = drive.get("role",       "Role")
    budget     = drive.get("budget_ctc", 0)
    date_str   = datetime.now().strftime("%d %B %Y")

    total        = len(candidates)
    selected_cnt = len([c for c in candidates if c.get("status") == "Selected"])
    avg_score    = sum(c.get("match_score", 0) for c in candidates) / max(total, 1)
    sorted_cands = sorted(candidates, key=lambda x: x.get("match_score", 0), reverse=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 35

    # Title
    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value     = "AI RECRUITMENT REPORT"
    title_cell.fill      = _fill(INDIGO)
    title_cell.font      = _font(bold=True, color=WHITE, size=16)
    title_cell.alignment = _align(h="center")
    ws1.row_dimensions[1].height = 36

    # Drive info
    info_rows = [
        ("Drive Name",    drive_name),
        ("Company",       company),
        ("Role",          role),
        ("Budget CTC",    f"{budget} LPA"),
        ("Generated On",  date_str),
    ]
    for i, (label, value) in enumerate(info_rows, start=2):
        ws1.cell(row=i, column=1, value=label).font  = _font(bold=True, color=DARK)
        ws1.cell(row=i, column=1).fill               = _fill(LIGHT_BG)
        ws1.cell(row=i, column=1).alignment          = _align()
        ws1.cell(row=i, column=1).border             = _border()
        ws1.cell(row=i, column=2, value=value).font  = _font(color="111827")
        ws1.cell(row=i, column=2).fill               = _fill(WHITE)
        ws1.cell(row=i, column=2).alignment          = _align()
        ws1.cell(row=i, column=2).border             = _border()

    ws1.append([])

    # Stats header
    stats_row = len(info_rows) + 3
    stats_labels = ["Total Candidates", "Selected", "Avg Match Score", "Budget CTC"]
    stats_values = [str(total), str(selected_cnt), f"{avg_score:.1f}%", f"{budget} LPA"]

    ws1.merge_cells(f"A{stats_row}:D{stats_row}")
    ws1[f"A{stats_row}"].value     = "RECRUITMENT STATISTICS"
    ws1[f"A{stats_row}"].fill      = _fill(VIOLET)
    ws1[f"A{stats_row}"].font      = _font(bold=True, color=WHITE, size=12)
    ws1[f"A{stats_row}"].alignment = _align(h="center")
    ws1.row_dimensions[stats_row].height = 28

    for col, (label, value) in enumerate(zip(stats_labels, stats_values), 1):
        ws1.column_dimensions[get_column_letter(col)].width = 22
        lc = ws1.cell(row=stats_row+1, column=col, value=label)
        lc.fill = _fill(INDIGO); lc.font = _font(bold=True, color=WHITE, size=9)
        lc.alignment = _align(h="center"); lc.border = _border()

        vc = ws1.cell(row=stats_row+2, column=col, value=value)
        vc.fill = _fill(LIGHT_BG); vc.font = _font(bold=True, color=DARK, size=14)
        vc.alignment = _align(h="center"); vc.border = _border()
        ws1.row_dimensions[stats_row+2].height = 30

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — All Candidates
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Candidates")
    ws2.sheet_view.showGridLines = False

    headers = [
        "Rank", "Name", "Email", "Phone",
        "Match Score", "Status",
        "Experience (yrs)", "Notice Period (days)",
        "Current CTC (LPA)", "Expected CTC (LPA)",
        "Skills", "Education"
    ]
    col_widths = [6, 22, 28, 16, 13, 16, 18, 20, 18, 18, 50, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill      = _fill(INDIGO)
        cell.font      = _font(bold=True, color=WHITE, size=10)
        cell.alignment = _align(h="center")
        cell.border    = _border()
    ws2.row_dimensions[1].height = 24
    ws2.freeze_panes = "A2"

    for row_i, c in enumerate(sorted_cands, start=2):
        status = c.get("status", "Pending")
        score  = c.get("match_score", 0)

        if status == "Selected":
            row_fill = "ECFDF5"
        elif status == "Rejected":
            row_fill = "FEF2F2"
        elif row_i % 2 == 0:
            row_fill = LIGHT_BG
        else:
            row_fill = WHITE

        score_color = SUCCESS if score >= 75 else (WARNING if score >= 50 else DANGER)

        edu = c.get("education", [])
        edu_str = "; ".join(
            f"{e.get('degree','')} {e.get('institution','')}" for e in edu
        ) if edu else ""

        row_vals = [
            c.get("rank", ""),
            c.get("name", ""),
            c.get("email", ""),
            c.get("phone", ""),
            f"{score:.1f}%",
            status,
            c.get("experience_years", 0),
            c.get("notice_period", 90),
            c.get("current_ctc", 0),
            c.get("expected_ctc", 0),
            ", ".join(c.get("skills", [])),
            edu_str,
        ]

        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=row_i, column=col, value=val)
            cell.border    = _border()
            cell.alignment = _align(h="center" if col in [1,5,6,7,8,9,10] else "left", wrap=True)

            # Score column coloring
            if col == 5:
                cell.fill = _fill(score_color)
                cell.font = _font(bold=True, color=WHITE, size=10)
            elif col == 6:
                sc = SUCCESS if status == "Selected" else (DANGER if status == "Rejected" else WARNING)
                cell.fill = _fill(sc)
                cell.font = _font(bold=True, color=WHITE, size=10)
            else:
                cell.fill = _fill(row_fill)
                cell.font = _font(color="111827", size=10)

        ws2.row_dimensions[row_i].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Selected Candidates
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Selected Candidates")
    ws3.sheet_view.showGridLines = False

    sel_headers = ["Name", "Score", "Experience", "Current CTC", "Expected CTC",
                   "Notice Period", "Skills", "Education"]
    sel_widths  = [22, 12, 14, 16, 16, 16, 50, 30]

    for col, (h, w) in enumerate(zip(sel_headers, sel_widths), 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = _fill(SUCCESS); cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align(h="center"); cell.border = _border()
    ws3.row_dimensions[1].height = 24
    ws3.freeze_panes = "A2"

    selected_list = [c for c in sorted_cands if c.get("status") == "Selected"]
    if not selected_list:
        ws3.cell(row=2, column=1, value="No candidates selected yet.")
    else:
        for row_i, c in enumerate(selected_list, start=2):
            edu = c.get("education", [])
            edu_str = "; ".join(f"{e.get('degree','')} {e.get('institution','')}" for e in edu)
            row_fill = LIGHT_BG if row_i % 2 == 0 else "ECFDF5"
            vals = [
                c.get("name",""), f"{c.get('match_score',0):.1f}%",
                f"{c.get('experience_years',0)} yrs",
                f"{c.get('current_ctc',0)} LPA", f"{c.get('expected_ctc',0)} LPA",
                f"{c.get('notice_period',90)} days",
                ", ".join(c.get("skills",[])), edu_str,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws3.cell(row=row_i, column=col, value=val)
                cell.fill = _fill(row_fill); cell.font = _font()
                cell.alignment = _align(h="center" if col in [2,3,4,5,6] else "left", wrap=True)
                cell.border = _border()
            ws3.row_dimensions[row_i].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — JD Analysis
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("JD Analysis")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 60

    ws4.merge_cells("A1:B1")
    ws4["A1"].value     = "Job Description Analysis"
    ws4["A1"].fill      = _fill(INDIGO)
    ws4["A1"].font      = _font(bold=True, color=WHITE, size=13)
    ws4["A1"].alignment = _align(h="center")
    ws4.row_dimensions[1].height = 30

    exp = jd_parsed.get("experience", {})
    jd_rows = [
        ("Role",               role),
        ("Company",            company),
        ("Experience",         f"{exp.get('min',0)} – {exp.get('max',5)} years"),
        ("Budget CTC",         f"{budget} LPA"),
        ("Required Skills",    ", ".join(jd_parsed.get("required_skills",  []))),
        ("Preferred Skills",   ", ".join(jd_parsed.get("preferred_skills", []))),
    ]
    for i, (label, value) in enumerate(jd_rows, start=2):
        lc = ws4.cell(row=i, column=1, value=label)
        lc.fill = _fill(LIGHT_BG); lc.font = _font(bold=True, color=DARK)
        lc.alignment = _align(); lc.border = _border()

        vc = ws4.cell(row=i, column=2, value=value)
        vc.fill = _fill(WHITE if i % 2 == 0 else LIGHT_GRAY)
        vc.font = _font(); vc.alignment = _align(wrap=True); vc.border = _border()
        ws4.row_dimensions[i].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Score Breakdown
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Score Breakdown")
    ws5.sheet_view.showGridLines = False

    sb_headers = ["Rank", "Name", "Total Score", "Skill Match",
                  "Experience", "Education", "Notice Period", "CTC Fit"]
    sb_widths  = [6, 22, 13, 13, 13, 13, 15, 12]

    for col, (h, w) in enumerate(zip(sb_headers, sb_widths), 1):
        ws5.column_dimensions[get_column_letter(col)].width = w
        cell = ws5.cell(row=1, column=col, value=h)
        cell.fill = _fill(VIOLET); cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align(h="center"); cell.border = _border()
    ws5.row_dimensions[1].height = 24
    ws5.freeze_panes = "A2"

    for row_i, c in enumerate(sorted_cands, start=2):
        bd = c.get("score_breakdown", {})
        score = c.get("match_score", 0)
        row_fill = LIGHT_BG if row_i % 2 == 0 else WHITE
        score_fill = SUCCESS if score >= 75 else (WARNING if score >= 50 else DANGER)

        vals = [
            c.get("rank",""),
            c.get("name",""),
            f"{score:.1f}%",
            f"{bd.get('skill_score',0):.1f}%",
            f"{bd.get('experience_score',0):.1f}%",
            f"{bd.get('education_score',0):.1f}%",
            f"{bd.get('notice_score',0):.1f}%",
            f"{bd.get('ctc_score',0):.1f}%",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws5.cell(row=row_i, column=col, value=val)
            cell.fill = _fill(score_fill if col == 3 else row_fill)
            cell.font = _font(bold=(col==3), color=WHITE if col==3 else "111827")
            cell.alignment = _align(h="center" if col != 2 else "left")
            cell.border = _border()
        ws5.row_dimensions[row_i].height = 18

    # ─── Save ─────────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()